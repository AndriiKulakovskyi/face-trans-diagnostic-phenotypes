#!/usr/bin/env python3
"""Populate DR-cohort source columns for NEUROPSYCHOLOGIE in the variable dictionary.

The new DR extraction recovers the neuropsychology block that the dictionary's
``DR column in CSV`` field never recorded (it was blank for all 133 NEUROPSYCHOLOGIE
rows, so ``Variable.source_col('DR')`` returned ``None`` and the loader surfaced zero DR
cognition — even though ``data/depression.csv`` already holds the items). This script
(idempotent / re-runnable) maps each DR-available neuropsych variable to its bare
lowercase DR CSV column, refreshes the now-stale ``Cluster readiness`` tier from the
actual per-cohort data presence, and writes an audit table for review.

Two mapping sources:
  * CORE — the dictionary's own ``Variable DR`` thesaurus name, lowercased (e.g.
    ``TMTA01`` → ``tmta01``). 59 rows have one.
  * ASSERTED — instruments whose DR column the dictionary omitted but whose name is
    identical across cohorts. Verbal fluency ``fv01``–``fv07`` (DR ~54% coverage) is the
    only such case found; including it makes the ``fluency`` construct trans-diagnostic.
    Disable with --no-fluency.

Every candidate is validated against the live ``depression.csv`` header and DR V0
coverage; anything missing or below --floor is left unmapped. ``Cluster readiness`` is
re-tiered ONLY when the data-presence tier changes (READY=3 cohorts / PARTIAL=2 /
NOT USABLE<=1), to minimise churn. Non-NEUROPSYCHOLOGIE rows are never touched.

Run:  python3 scripts/build_dr_neuropsych_mapping.py [--floor 0.30] [--no-fluency] [--dry-run]
Writes in place: data/face-common-vars.xlsx ; audit: results/dr_neuropsych_mapping_audit.csv
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "src"))

from trans_diag.domains import instrument_stem  # noqa: E402

XLSX = REPO_ROOT / "data" / "face-common-vars.xlsx"
DEP_CSV = REPO_ROOT / "data" / "depression.csv"
BP_CSV = REPO_ROOT / "data" / "bipolar.csv"
SZ_CSV = REPO_ROOT / "data" / "schizophrenia.csv"
AUDIT = REPO_ROOT / "results" / "dr_neuropsych_mapping_audit.csv"
SECTION = "NEUROPSYCHOLOGIE"

H_SECTION = "Section"
H_CANON = "Canonical name (merged single-cohort)"
H_VDR = "Variable DR"
H_BP = "BP column in CSV"
H_SZ = "SZ column in CSV"
H_DR = "DR column in CSV"
H_READY = "Cluster readiness"

# Instruments whose DR column the dictionary's "Variable DR" omitted but whose bare name
# is identical across cohorts. canonical -> DR csv column. Validated before use.
FLUENCY_ASSERTED = {f"fv0{i}": f"fv0{i}" for i in range(1, 8)}

# Curated cognitive constructs (mirrors the soon-to-be-removed scripts/14); used only to
# annotate the audit with each variable's construct and its trans-diagnostic status.
COGNITIVE_DOMAINS = {
    "memory_cvlt": ["cvlt"],
    "executive_tmt": ["tmtb", "tmtba"],
    "proc_speed": ["tmta", "code01_wais", "code02_wais", "code03_wais", "code04_wais",
                   "code05_wais", "ivt01_wais", "ivt02_wais", "ivt04_wais"],
    "working_memory": ["nbrut_w", "nstand_w", "vstand_w", "mcod_w", "mcoi_w", "mcoc_w",
                       "mcodemp_w", "mcoiemp_w", "empdid_w", "wais_mc_end_std_wais",
                       "wais_mc_env_std_wais", "wais_mc_cro_wais", "wais_mc_cro_std_wais"],
    "verbal_reasoning": ["similtot_wais", "similstd_wais", "similcr_wais"],
    "percept_reasoning": ["mat_tot_w", "mat_std_w", "mat_cr_w"],
    "fluency": ["fv"],
}
_MEMBER_TO_CONSTRUCT = {m: c for c, ms in COGNITIVE_DOMAINS.items() for m in ms}


def _v0_coverage(csv_path: Path) -> dict[str, float]:
    """Per-column V0 non-null fraction for a cohort CSV."""
    df = pd.read_csv(csv_path, low_memory=False, encoding="utf-8-sig")
    v0 = df[df["visit"].astype(str) == "V0"]
    return {c: float(v0[c].notna().mean()) for c in v0.columns}


def _tier_prefix(n_cohorts: int) -> str:
    return {3: "READY", 2: "PARTIAL"}.get(n_cohorts, "NOT USABLE")


def _tier_label(n_cohorts: int) -> str:
    return {
        3: "READY — Comparable in 3 cohorts (Tier A); data present in 3 CSVs "
           "(DR neuropsych recovered 2026-05)",
        2: "PARTIAL — Comparable in 2 of 3 cohorts (Tier B); data present in 2 CSVs",
    }.get(n_cohorts, "NOT USABLE — data present in only 1 CSV")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--floor", type=float, default=0.30,
                    help="min DR V0 coverage to accept a mapping (default 0.30, matches the pipeline)")
    ap.add_argument("--no-fluency", action="store_true",
                    help="skip the asserted fv01-07 fluency mapping")
    ap.add_argument("--dry-run", action="store_true",
                    help="compute + write the audit CSV but do NOT modify the xlsx")
    args = ap.parse_args()

    dep_cov = _v0_coverage(DEP_CSV)
    bp_cov = _v0_coverage(BP_CSV)
    sz_cov = _v0_coverage(SZ_CSV)
    dep_cols = set(dep_cov)

    df = pd.read_excel(XLSX, sheet_name="Sheet1")
    nps = df[df[H_SECTION] == SECTION]
    print(f"{len(nps)} NEUROPSYCHOLOGIE rows; floor={args.floor}; "
          f"fluency={'OFF' if args.no_fluency else 'ON'}; "
          f"mode={'DRY-RUN' if args.dry_run else 'APPLY'}")

    # Resolve the intended DR column per canonical (CORE then ASSERTED), validate it.
    planned: dict[str, dict] = {}      # canonical -> {dr_col, source, cov_dr}
    for _, r in nps.iterrows():
        canon = str(r[H_CANON]).strip()
        vdr = r[H_VDR]
        cand, source = None, None
        if pd.notna(vdr):
            cand, source = str(vdr).strip().lower(), "Variable DR"
        elif not args.no_fluency and canon in FLUENCY_ASSERTED:
            cand, source = FLUENCY_ASSERTED[canon], "asserted (fv name identity)"
        if cand is None:
            continue
        cov = dep_cov.get(cand, float("nan"))
        ok = (cand in dep_cols) and pd.notna(cov) and (cov >= args.floor)
        planned[canon] = {"dr_col": cand if ok else None, "source": source,
                          "cov_dr": cov, "exists": cand in dep_cols, "accepted": ok}

    # Build the audit + decide readiness re-tiering (uses the *planned* DR presence).
    audit_rows: list[dict] = []
    readiness_updates: dict[str, str] = {}   # canonical -> new label
    for _, r in nps.iterrows():
        canon = str(r[H_CANON]).strip()
        stem = instrument_stem(canon)
        construct = _MEMBER_TO_CONSTRUCT.get(stem, "")
        bp_col = None if pd.isna(r[H_BP]) else str(r[H_BP]).strip()
        sz_col = None if pd.isna(r[H_SZ]) else str(r[H_SZ]).strip()
        p = planned.get(canon, {})
        dr_col = p.get("dr_col")
        n_cohorts = sum(x is not None for x in (bp_col, sz_col, dr_col))
        old_ready = "" if pd.isna(r[H_READY]) else str(r[H_READY]).strip()
        new_ready = old_ready
        if not old_ready.startswith(_tier_prefix(n_cohorts)):
            new_ready = _tier_label(n_cohorts)
            readiness_updates[canon] = new_ready
        audit_rows.append({
            "canonical": canon, "construct": construct, "stem": stem,
            "variable_dr": "" if pd.isna(r[H_VDR]) else str(r[H_VDR]).strip(),
            "dr_col_final": dr_col or "", "map_source": p.get("source", ""),
            "accepted": bool(p.get("accepted", False)),
            "dr_col_exists": bool(p.get("exists", False)),
            "cov_dr_v0": round(p.get("cov_dr", float("nan")), 3) if p else float("nan"),
            "cov_bp_v0": round(bp_cov.get(bp_col, float("nan")), 3) if bp_col else float("nan"),
            "cov_sz_v0": round(sz_cov.get(sz_col, float("nan")), 3) if sz_col else float("nan"),
            "bp_col": bp_col or "", "sz_col": sz_col or "", "n_cohorts": n_cohorts,
            "readiness_old": old_ready, "readiness_new": new_ready,
            "readiness_changed": new_ready != old_ready,
        })

    n_map = sum(1 for v in planned.values() if v["dr_col"])
    n_reject = sum(1 for v in planned.values() if not v["dr_col"])
    AUDIT.parent.mkdir(exist_ok=True)
    pd.DataFrame(audit_rows).to_csv(AUDIT, index=False)
    print(f"mapped DR columns: {n_map}  | rejected (missing/low-cov): {n_reject}  "
          f"| readiness re-tiered: {len(readiness_updates)}")
    print(f"audit -> {AUDIT.relative_to(REPO_ROOT)}")

    # Construct-level trans-diagnostic summary
    print("\nconstruct -> cohorts with an ACCEPTED member (locked-model admissibility):")
    for con in COGNITIVE_DOMAINS:
        rows = [a for a in audit_rows if a["construct"] == con]
        dr = sorted({a["stem"] for a in rows if a["accepted"]})
        bp = sorted({a["stem"] for a in rows if a["bp_col"]})
        tag = "DR-PRESENT -> trans-diagnostic" if dr else "BP/SZ-only -> excluded from locked model"
        print(f"  {con:18s} BP={len(bp)} DR={len(dr)}  [{tag}]  DR stems={dr}")

    if args.dry_run:
        print("\nDRY-RUN: xlsx unchanged.")
        return 0

    # Apply to the xlsx in place (openpyxl, cell-level — preserves all other content).
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Sheet1"]
    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    for col in (H_SECTION, H_CANON, H_DR, H_READY):
        if col not in header:
            raise SystemExit(f"header {col!r} not found in Sheet1")
    n_dr_writes = n_ready_writes = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=header[H_SECTION]).value != SECTION:
            continue
        canon = str(ws.cell(row=row, column=header[H_CANON]).value).strip()
        if canon in planned and planned[canon]["dr_col"]:
            ws.cell(row=row, column=header[H_DR]).value = planned[canon]["dr_col"]
            n_dr_writes += 1
        if canon in readiness_updates:
            ws.cell(row=row, column=header[H_READY]).value = readiness_updates[canon]
            n_ready_writes += 1
    wb.save(XLSX)
    print(f"\nAPPLIED: wrote {n_dr_writes} DR columns + {n_ready_writes} readiness labels "
          f"to {XLSX.relative_to(REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
